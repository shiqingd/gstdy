#!/usr/bin/env python3

import os
import sys
import sh
import argparse
import textwrap
import traceback
import re
import json
from datetime import datetime
import time
from io import StringIO

if not "DJANGO_SETTINGS_MODULE" in os.environ:
		os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings.settings")
		import django
		django.setup()

from framework.models import *

import lib.dry_run
from lib.pushd import pushd
from lib.colortext import ANSIColor, RGBColor

from django.db.models import F, Q

def __intel_authored(author):
	'''
	Test if a string from a commit message is an Intel email address
	'''
	try:
		return re.search("@(?:linux\.){0,1}intel.com$", author)
	except:
		return False

def __is_email_address(text):
	'''
	Test if a string is an valid email address
	'''
	regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
	# and the string in search() method 
	return re.match(regex,text)


class DebtPatch(object):
	def __init__(self, **kwargs):
		self.__dict__.update(kwargs)

	def __repr__(self):
		return self.__dict__

	def __str__(self):
		return str(self.__dict__)

def __do_reports(args):
	'''
	Update the staging commits for the specified kernel
	'''
	for kernel_name in args.kernel:
		kernel = Kernel.objects.get(name = kernel_name)
		# FIXME monkey patch - don't call save() on this object!
		kernel.base_kernel = 'v'+kernel.base_kernel

		krs = KernelRepoSet.objects.filter(kernel = kernel, repo__repotype__repotype = 'src')
		for kr in krs:
			print(RGBColor(255,165,0, "\tProcessing Repo {}".format(kr.repo.project)))
			__git = kr.repo.initialize(scmdir=os.path.join(os.environ["WORKSPACE"],kr.repo.project))	# returns a sh.Command() object
			with pushd (kr.repo.scmdir):
				# Use each tracking/staging branch configured for this kernel in each repo
				tbs = TrackerBranch.objects.filter(kernel = kernel, repo = kr.repo)
				for tb in tbs:
					try:
						__git("checkout" , tb.branch)
					except Exception as e:
						print("Cannot checkout {}: {}".format(e.args[0].strip().split('\n')[-1]))
						continue
					revlist = __git("rev-list", "--no-merges", "--reverse" , "{}..{}".format(kernel.current_baseline, tb.branch).strip())
					revlist = revlist.split()
					print(ANSIColor("yellow", "\t\tBranch Window: %s..%s : %d patches" % (kernel.current_baseline, tb.branch, len(revlist))))
					print(ANSIColor("bright_blue", "Processing Hashes..."))
					debt = []
					i = 0
					for rev in revlist:
						patch_body = __git("show" , "-1", "--format=%b" , rev, _tty_out=False).stdout.decode(errors='ignore')
						payload_hash = __git("patch-id", _in=StringIO(patch_body), _tty_out=False).split()[0]
						patch = StableKernel.objects.filter(payload_hash = payload_hash).first()
						if not patch:
							iii = __git("log", "-1", "--format=%s\n%ae\n%ct", rev, _tty_out=False).stdout.decode(errors='ignore').strip().split('\n')
							subject = iii[0]
							author = iii[1]
							files = __git("show", "-1", "--format=", "--name-only", rev, _tty_out=False).stdout.decode(errors='ignore').strip().split('\n')
							if not __intel_authored(author):
								print (ANSIColor("cyan", rev, payload_hash, "NON-INTEL-AUTHORED", author))
								i += 1
								if not (i % 100):
									print(i)
								continue
							dp = DebtPatch( commit_id = rev, subject = subject, author = author, files = files, upstream_commit = '', upstream_author = '', upstream_tag = '')
							patch = StableKernel.objects.filter(subject__startswith = subject ).first()
							if patch:
								print (ANSIColor("yellow", rev, "MAY BE DERIVED FROM", patch.subject, "commit", patch.commit_id , "tag", patch.tag))
								dp.upstream_commit = patch.commit_id
								dp.upstream_author = patch.author
								dp.subject = patch.subject  # in case of quilt truncation
								dp.upstream_tag = patch.tag
							else:
								print (ANSIColor("magenta", rev, payload_hash, author, "NOT UPSTREAM"))
							debt.append(dp)
						else:
#							print(ANSIColor("green", payload_hash, "UPSTREAM at" , patch.tag, patch.author, patch.commit_id))
							pass
						i += 1
						if not (i % 100):
							print(i)
					for dp in debt:
						if TDWhitelist.objects.filter(commit_id = dp.commit_id):
							print("INFO: patch {} ({}) WhiteListed".format(dp.commit_id, dp.subject))
							debt.remove(dp)
					print(len(debt), "patches have no upstream commit")
					if args.report:
						__compute_domain_report_xlsx(kernel, debt, tb.branch)

def __size_columns(ws):
	for column_cells in ws.columns:
		cl = column_cells[0].column_letter
		new_width = 0.0
		for cell in column_cells:
			if cell.value is not None:
				new_width = max(new_width, len(str(cell.value)))
		new_width = (new_width + 2) * 1.2
		ws.column_dimensions[cl].width = new_width
		print(str(ws), "Width", cl, "=", new_width)

def __compute_domain_report_xlsx(kernel, debt, staging_branch):
	'''
	Produce a Microsoft Excel 201x workbook (.XLSX) file
	breaking down Technical Debt By Domain
	'''
	from openpyxl import Workbook
	from openpyxl.styles import Font
	from openpyxl.styles.fonts import DEFAULT_FONT

	dlist = {}
	dlist_count = {}
	wb = Workbook()
	# NOTE: using a monospace font helps to properly auto-size the columns
	DEFAULT_FONT.name = 'Courier New'

	print("Calculating Debt By Domain ... ")

	for dp in debt:
		try:
			# Handle corner cases where patches that normally belong
			# in one domain get patched by a differenct domain
			rm = TDRemap.objects.get(commit_id = dp.commit_id)
			patchdomains = [ rm.domain ]
		except:
			patchdomains = __compute_domains(dp.files)
		for domain in patchdomains:
			if not domain in dlist.keys():
				dlist[domain] = []
			dlist[domain].append(dp)
	for dd in dlist.keys():
		if dd.rfu_0 & 1:
			continue
		print(dd.name)
		ws = wb.create_sheet(dd.name)
		row=1
		column=1
		ws.cell(row=row,column=column, value="DOMAIN")
		column += 1
		ws.cell(row=row,column=column, value="COMMIT ID")
		column += 1
		ws.cell(row=row,column=column, value="SUBJECT")
		column += 1
		ws.cell(row=row,column=column, value="AUTHOR")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM COMMIT")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM AUTHOR")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM LOCATION")
		column += 1
		ws.cell(row=row,column=column, value="EXEMPTED BY")
		column += 1
		ws.cell(row=row,column=column, value="EXEMPTION REASON")
#       column += 1
#       ws.cell(row=row,column=column, value="FILES")
		column = 1
		r = ws.row_dimensions[row]
		r.font = Font(bold=True)
		row += 1
		dlist_count[dd] = 0
		for dp in dlist[dd]:
			ws.cell(row=row,column=column, value=dd.name)
			column += 1
			ws.cell(row=row,column=column, value=dp.commit_id)
			column += 1
			ws.cell(row=row,column=column, value=dp.subject)
			column += 1
			ws.cell(row=row,column=column, value=dp.author)
			column += 1
			ws.cell(row=row,column=column, value=dp.upstream_commit)
			column += 1
			ws.cell(row=row,column=column, value=dp.upstream_author)
			column += 1
			ws.cell(row=row,column=column, value=dp.upstream_tag)
			column += 1
			ws.cell(row=row,column=column, value="")
			column += 1
			ws.cell(row=row,column=column, value="")
#			column += 1
#			for f in dp.files:
#				ws.cell(row=row,column=column, value=f)
#				column += 1
			row += 1
			column = 1
			if not dp.upstream_commit:
				dlist_count[dd] += 1
		__size_columns(ws)

	ws = wb.create_sheet('Summary', 0)
	row=2
	column=1
	ws.cell(row=row,column=column, value="Kernel Technical Debt Summary for Merge Window {}".format(kernel.merge_baseline))
	row += 1
	ws.cell(row=row,column=column, value="Staging Branch: {}".format(staging_branch))
	row += 2
	column = 1
	ws.cell(row=row,column=column, value="DOMAIN")
	column += 1
	ws.cell(row=row,column=column, value="COUNT")
	column = 1
	r = ws.row_dimensions[row]
	r.font = Font(bold=True)
	row += 1
	sumstartrow = row
	for dd in dlist.keys():
		if dd.rfu_0 & 1:
			continue
		ws.cell(row=row,column=column, value=dd.name)
		column += 1
		ws.cell(row=row,column=column, value=dlist_count[dd])
		row += 1
		column = 1
	__size_columns(ws)

	sumendrow = row-1
	row += 1
	ws.cell(row=row,column=column, value="TOTAL")
	column += 1
	ws.cell(row=row,column=column, value="=SUM(B"+str(sumstartrow)+":B"+str(sumendrow))
	# Delete Default sheet named 'Sheet'
	del wb['Sheet']
	# Sort Sheets alphabetically by domain
	wb._sheets.sort(key=lambda ws: ws.title)
	# Move Summary sheet to front of sheet list
	summary_sheet = wb._sheets.pop(wb._sheets.index(wb['Summary']))
	wb._sheets.insert(0, summary_sheet)
	# Save Workbook
	printable_branch_name = re.sub('[\.\-\/]', '_', staging_branch)
	output_file = "/tmp/{}_{}_techdebt_{}.xlsx".format(kernel.current_baseline, printable_branch_name, datetime.fromtimestamp(time.time()).strftime('%m%d'))
	wb.save(output_file)
	print("Results saved to {}".format(output_file))

missing_doms = None
def __compute_domains(files):
	'''
	Determine the appropriate domain for a patch based on files modified
	'''
	global missing_doms
	g_paths =  PathToDomain.objects.all();
	g_regexps = {}
	for p in g_paths:
		g_regexps[p.pattern] = re.compile(p.pattern)
	domainlist = []
	for f in files:
		matchlist = []
		for p in g_paths:
			match = g_regexps[p.pattern].search(f)
			if match:
#				print(ANSIColor("yellow", match.group(0)))
				matchlist.append( ( p , match ) )
		if len(matchlist) == 0:
			print('FILE: <'+f+'> NO DOMAIN FOUND')
			# log any patches files for which a reasonable domain cannot be determined
			if not missing_doms:
				missing_doms = open("/tmp/missing_doms.sql", "w+")
			missing_doms.write("INSERT into framework_pathtodomain (pattern, domain_id) VALUES ('{}', 999) ; \n".format(f))
			continue
#		print('FILE:', '<'+f+'>')
		chosenmatch = matchlist[0]
		if len(matchlist) > 1:
#			print('FILE:', '<'+f+'>')
			# Try to find a "best fit" based on match string 
			matchdelta = 9999
			for match in matchlist:
				delta = len(f) - len(match[1].group(0))
				if delta < matchdelta:
					chosenmatch = match
					matchdelta = delta
#				print(ANSIColor("green", '	', match[0].pattern, match[1].group(0), match[0].domain.name))
#		print(ANSIColor("white", 'WINNER:	', chosenmatch[0].pattern, chosenmatch[1].group(0), chosenmatch[0].domain.name))
		if chosenmatch[0].domain not in domainlist:
			domainlist.append(chosenmatch[0].domain)
		if len(domainlist) > 1:
			# Filter out domains that we don't want to double-domain a patch with
			no_double_domains = Domain.objects.annotate(dbl_filter=F('rfu_0').bitand(2)).filter(dbl_filter = 2)
			for dom in domainlist:
				if dom in no_double_domains:
					domainlist.remove(dom)


	return domainlist

def __update_whitelist(xlsx_file):
	'''
	Read an Excel Workbook and add to the TDWhitelist table
	(currently this method is run from a Python console)
	'''
	from openpyxl import load_workbook
	wb = load_workbook(filename = xlsx_file)
	for ws in wb.worksheets[1:]:
		for cell in ws['H'][1:]:
				if __is_email_address(str(cell.value)):
					index = cell.row-1
					commit_id = ws['B'][index].value
					upstream_tag = ws['G'][index].value
					if upstream_tag is None:
						upstream_tag = ''
					whitelisted_by = ws['H'][index].value
					reason = ws['I'][index].value
					wl = TDWhitelist(commit_id = commit_id, upstream_tag = upstream_tag, whitelisted_by = whitelisted_by, reason = reason)
					print(repr(wl))
					wl.save()

if __name__ == '__main__':
	parser = argparse.ArgumentParser(prog=sys.argv[0])
	parser.add_argument('--kernel', '-k', type=Kernel.validate, help="Valid Values: "+Kernel.list()+" default: all kernels in DB")
	parser.add_argument('--report', '-R', action='store_true', default=False, help="generate report(s)")
	args = parser.parse_args()

	assert("WORKSPACE" in os.environ)

	if not args.kernel:
		args.kernel = json.loads(re.sub('\'','"',Kernel.list()))		# produce TD report for all supported kernels
	else:
		args.kernel = [ args.kernel ]
	print(args.kernel)
	__do_reports(args)
	print('Done')
