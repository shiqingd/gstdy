#!/usr/bin/env python3

import os
import sys
import sh
import argparse
import textwrap
import traceback
import re
import json
from datetime import datetime
import time

if not "DJANGO_SETTINGS_MODULE" in os.environ:
        os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings.settings")
        import django
        django.setup()

from framework.models import *

import lib.dry_run
from lib.pushd import pushd
from lib.colortext import ANSIColor, RGBColor
from lib.PatchParser import get_patch_id, get_patch_dict

from django.db.models import F, Q

def __intel_authored(author):
	'''
	Test if a string from a commit message is an Intel email address
	'''
	try:
		return re.search("@(?:linux\.){0,1}intel.com$", author[0]["email"])
	except:
		return False

def __is_email_address(text):
	'''
	Test if a string is an valid email address
	'''
	regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
	# and the string in search() method 
	return re.match(regex,text)


def __scan(args):
	'''
	Scan the Technical Debt for the kernel(s) specified in args
	'''
	for kernel_name in args.kernel:
		kernel = Kernel.objects.get(name = kernel_name)
		# FIXME monkey patch - don't call save() on this object!
		kernel.base_kernel = 'v'+kernel.base_kernel

		print(ANSIColor("cyan", '"Processing Tech Debt for Kernel {}, Base Kernel {}, Domain Baseline {}, Merge Tip {}"'.format(kernel.name, kernel.base_kernel, kernel.current_baseline, kernel.merge_baseline)))
		krs = KernelRepoSet.objects.filter(kernel = kernel, repo__repotype__repotype = 'src')
		for kr in krs:
			print(RGBColor(255,165,0, "\tProcessing Repo {}".format(kr.repo.project)))
			__git = kr.repo.initialize(scmdir=os.path.join(os.environ["WORKSPACE"],kr.repo.project))	# returns a sh.Command() object
			# Compute using patch-id hash (more reliable)
			with pushd (kr.repo.scmdir):
				# Use each tracking/staging branch configured for this kernel in each repo
				tbs = TrackerBranch.objects.filter(kernel = kernel, repo = kr.repo)
				for tb in tbs:
					try:
						__git("checkout" , tb.branch)
					except Exception as e:
						print("Cannot checkout {}: {}".format(e.args[0].strip().split('\n')[-1]))
						continue
					revlist = __git("rev-list", "--no-merges", "--reverse" , "{}..{}".format(kernel.current_baseline, tb.branch).strip())
					revlist = revlist.split()
					print(ANSIColor("yellow", "\t\tBranch Window: %s..%s : %d patches" % (kernel.current_baseline, tb.branch, len(revlist))))
					print(ANSIColor("bright_blue", "Processing Hashes..."))
					i = 0
					debt_candidates = {}
					for rev in revlist:
						debt_candidates[rev] = get_patch_dict(rev) # get patch details
						i += 1
						if (i % 100 == 0) : print(i)

					revlist = __git("rev-list", "--no-merges", "--reverse" , "{}..{}".format(kernel.base_kernel, kernel.merge_baseline).strip())
					revlist = revlist.split()
				print(ANSIColor("green", "\tUpdating Merge Window: %s..%s : %d patches" % (kernel.base_kernel, kernel.merge_baseline, len(revlist))))
				print(ANSIColor("bright_green", "Processing Hashes..."))
				i = 0
				upstream_patches = {}
				for rev in revlist:
					pd = get_patch_dict(rev)		# get patch details
					upstream_patches[pd["patch-id"]] = pd
					i += 1
					if (i % 100 == 0) : print(i)
				
				print("Calculating Debt ... ")
				the_debt = {}
				for k in debt_candidates.keys():
					if not debt_candidates[k]["patch-id"] in upstream_patches.keys() and __intel_authored(debt_candidates[k]["Author"]):
						# If the patch has not been whitelisted, add it to the debt
						wl = TDWhitelist.objects.filter(commit_id = debt_candidates[k]["Commit"][0])
						if not wl:
							the_debt[k] = debt_candidates[k]
				print("Debt Calculation Complete")
				j = { "UPSTREAM" : upstream_patches,  "CANDIDATES" : debt_candidates , "DEBT" : the_debt }
				__compute_domain_xlsx(j["DEBT"], j["UPSTREAM"], kernel, list(b["branch"] for b in list(tbs.values('branch'))))

def __compute_domain_csv(debt, upstreamed):
	'''
	OBSOLETE: Produce a CSV file
	breaking down Technical Debt By Domain
	'''
	dlist = {}
	for dp in debt.values():
		dp["upstream_origin"] = ''
		dp["upstream_author"] = ''
		for u in upstreamed.values():
			if dp["Subject"][0] == u["Subject"][0]:
				print(dp["Subject"][0],u["Subject"][0])
				dp["upstream_origin"] = upstreamed[dp]["Commit"][0]
				dp["upstream_author"] = upstreamed[dp]["Author"][0][email]
				break
		patchdomains = __compute_domains(dp["files"])
		for domain in patchdomains:
			if not domain in dlist.keys():
				dlist[domain] = []
			dlist[domain].append(dp)
	for dd in dlist.keys():
		for dp in dlist[dd]:
			print('"{}","{}","{}","{}","{}","{}",{}'.format(dd.name, dp["Commit"][0], dp["Subject"][0], dp["Author"][0]["email"], dp["upstream_origin"], dp["upstream_author"], ','.join(map(lambda x: '"'+x+'"', dp["files"]))))


def __compute_domain_xlsx(debt, upstreamed, kernel, staging_branches):
	'''
	Produce a Microsoft Excel 201x workbook (.XLSX) file
	breaking down Technical Debt By Domain
	'''
	from openpyxl import Workbook
	from openpyxl.styles import Font
	from openpyxl.styles.fonts import DEFAULT_FONT

	dlist = {}
	wb = Workbook()
	# NOTE: using a monospace font helps to properly auto-size the columns
	DEFAULT_FONT.name = 'Courier New'
	for dp in debt.values():
		dp["upstream_origin"] = ''
		dp["upstream_author"] = ''
		for u in upstreamed.values():
			if dp["Subject"][0] == u["Subject"][0]:
				print(dp["Subject"][0],u["Subject"][0])
				dp["upstream_origin"] = u["Commit"][0]
				dp["upstream_author"] = u["Author"][0]["email"]
				break
		try:
			# Handle corner cases where patches that normally belong
			# in one domain get patched by a differenct domain
			rm = TDRemap.objects.get(commit_id = dp["Commit"][0])
			patchdomains = [ rm.domain ]
		except:
			patchdomains = __compute_domains(dp["files"])
		for domain in patchdomains:
			if not domain in dlist.keys():
				dlist[domain] = []
			dlist[domain].append(dp)
	for dd in dlist.keys():
		if dd.rfu_0 & 1:
			continue
		print(dd.name)
		ws = wb.create_sheet(dd.name)
		row=1
		column=1
		ws.cell(row=row,column=column, value="DOMAIN")
		column += 1
		ws.cell(row=row,column=column, value="COMMIT ID")
		column += 1
		ws.cell(row=row,column=column, value="SUBJECT")
		column += 1
		ws.cell(row=row,column=column, value="AUTHOR")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM COMMIT")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM AUTHOR")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM LOCATION")
		column += 1
		ws.cell(row=row,column=column, value="WHITELISTED BY")
		column += 1
		ws.cell(row=row,column=column, value="WHITELIST REASON")
		column += 1
		ws.cell(row=row,column=column, value="FILES")
		column = 1
		r = ws.row_dimensions[row]
		r.font = Font(bold=True)
		row += 1
		for dp in dlist[dd]:
			ws.cell(row=row,column=column, value=dd.name)
			column += 1
			ws.cell(row=row,column=column, value=dp["Commit"][0])
			column += 1
			ws.cell(row=row,column=column, value=dp["Subject"][0])
			column += 1
			try:
				ws.cell(row=row,column=column, value=dp["Author"][0]["email"])
			except:
				try:
					print(dp["Author"])
					ws.cell(row=row,column=column, value=dp["Author"])
				except:
					print(dp)
					ws.cell(row=row,column=column, value='')
			column += 1
			ws.cell(row=row,column=column, value=dp["upstream_origin"])
			column += 1
			ws.cell(row=row,column=column, value=dp["upstream_author"])
			column += 1
			ws.cell(row=row,column=column, value="")
			column += 1
			ws.cell(row=row,column=column, value="")
			column += 1
			ws.cell(row=row,column=column, value="")
			column += 1
			for f in dp["files"]:
				ws.cell(row=row,column=column, value=f)
				column += 1
			row += 1
			column = 1
		for column_cells in ws.columns:
			length = max(len(cell.value is None and "" or str(cell.value)) for cell in column_cells)
			ws.column_dimensions[column_cells[0].column].width = length + 5
	ws = wb.create_sheet('Summary', 0)
	row=1
	column=1
	ws.cell(row=row,column=column, value="Kernel Technical Debt Summary for Merge Window {}".format(kernel.merge_baseline))
	row += 1
	ws.cell(row=row,column=column, value="Staging Branch(es): {}..".format(','.join(b for b in staging_branches)))
	row += 2
	column = 1
	ws.cell(row=row,column=column, value="DOMAIN")
	column += 1
	ws.cell(row=row,column=column, value="COUNT")
	column = 1
	r = ws.row_dimensions[row]
	r.font = Font(bold=True)
	row += 1
	for dd in dlist.keys():
		if dd.rfu_0 == 1:
			continue
		ws.cell(row=row,column=column, value=dd.name)
		column += 1
		ws.cell(row=row,column=column, value=len(dlist[dd]))
		row += 1
		column = 1
		for column_cells in ws.columns:
			length = max(len(cell.value is None and "" or str(cell.value)) for cell in column_cells)
			ws.column_dimensions[column_cells[0].column].width = length + 5
	# Delete Default sheet named 'Sheet'
	del wb['Sheet']
	# Sort Sheets alphabetically by domain
	wb._sheets.sort(key=lambda ws: ws.title)
	# Move Summary sheet to front of sheet list
	summary_sheet = wb._sheets.pop(wb._sheets.index(wb['Summary']))
	wb._sheets.insert(0, summary_sheet)
	# Save Workbook
	output_file = "/tmp/{}_techdebt_{}.xlsx".format(kernel.current_baseline, datetime.fromtimestamp(time.time()).strftime('%m%d'))
	wb.save(output_file)
	print("Results saved to {}".format(output_file))

missing_doms = None
def __compute_domains(files):
	'''
	Determine the appropriate domain for a patch based on files modified
	'''
	global missing_doms
	g_paths =  PathToDomain.objects.all();
	g_regexps = {}
	for p in g_paths:
		g_regexps[p.pattern] = re.compile(p.pattern)
	domainlist = []
	for f in files:
		matchlist = []
		for p in g_paths:
			match = g_regexps[p.pattern].search(f)
			if match:
#				print(ANSIColor("yellow", match.group(0)))
				matchlist.append( ( p , match ) )
		if len(matchlist) == 0:
			print('FILE: <'+f+'> NO DOMAIN FOUND')
			# log any patches files for which a reasonable domain cannot be determined
			if not missing_doms:
				missing_doms = open("/tmp/missing_doms.sql", "w+")
			missing_doms.write("INSERT into framework_pathtodomain (pattern, domain_id) VALUES ('{}', 999) ; \n".format(f))
			continue
#		print('FILE:', '<'+f+'>')
		chosenmatch = matchlist[0]
		if len(matchlist) > 1:
#			print('FILE:', '<'+f+'>')
			# Try to find a "best fit" based on match string 
			matchdelta = 9999
			for match in matchlist:
				delta = len(f) - len(match[1].group(0))
				if delta < matchdelta:
					chosenmatch = match
					matchdelta = delta
#				print(ANSIColor("green", '	', match[0].pattern, match[1].group(0), match[0].domain.name))
#		print(ANSIColor("white", 'WINNER:	', chosenmatch[0].pattern, chosenmatch[1].group(0), chosenmatch[0].domain.name))
		if chosenmatch[0].domain not in domainlist:
			domainlist.append(chosenmatch[0].domain)
		if len(domainlist) > 1:
			# Filter out domains that we don't want to double-domain a patch with
			no_double_domains = Domain.objects.annotate(dbl_filter=F('rfu_0').bitand(2)).filter(dbl_filter = 2)
			for dom in domainlist:
				if dom in no_double_domains:
					domainlist.remove(dom)


	return domainlist

def __update_whitelist(xlsx_file):
	'''
	Read an Excel Workbook and add to the TDWhitelist table
	(currently this method is run from a Python console)
	'''
	from openpyxl import load_workbook
	wb = load_workbook(filename = xlsx_file)
	for ws in wb.worksheets[1:]:
		for cell in ws['H'][1:]:
				if __is_email_address(str(cell.value)):
					index = cell.row-1
					commit_id = ws['B'][index].value
					upstream_tag = ws['G'][index].value
					if upstream_tag is None:
						upstream_tag = ''
					whitelisted_by = ws['H'][index].value
					reason = ws['I'][index].value
					wl = TDWhitelist(commit_id = commit_id, upstream_tag = upstream_tag, whitelisted_by = whitelisted_by, reason = reason)
					print(repr(wl))
					wl.save()

if __name__ == '__main__':
	parser = argparse.ArgumentParser(prog=sys.argv[0])
	parser.add_argument('--kernel', '-k', type=Kernel.validate, help="Valid Values: "+Kernel.list()+" default: all kernels in DB")
	args = parser.parse_args()

	assert("WORKSPACE" in os.environ)

	if not args.kernel:
		args.kernel = json.loads(re.sub('\'','"',Kernel.list()))		# produce TD report for all supported kernels
	else:
		args.kernel = [ args.kernel ]
	print(args.kernel)
	__scan(args)
	print('Done')
