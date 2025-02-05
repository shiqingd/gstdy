#!/usr/bin/env python3

import os
import sys
import sh
import argparse
import textwrap
import traceback
import re
import json
from datetime import datetime, timedelta
import time
from glob import glob
import shutil

if not "DJANGO_SETTINGS_MODULE" in os.environ:
		os.environ.setdefault("DJANGO_SETTINGS_MODULE", "settings.settings")
		import django
		django.setup()

from framework.models import *

import lib.dry_run
from lib.pushd import pushd
from lib.colortext import ANSIColor, RGBColor
from lib.PatchParser import get_patch_id, get_patch_dict
from lib.kversion import KernelVersion

from django.db.models import F, Q

import pytz
from django.utils import timezone

def __intel_authored(author):
	'''
	Test if a string from a commit message is an Intel email address
	'''
	try:
		return re.search("@(?:linux\.){0,1}intel.com$", author)
	except:
		return False

def __is_email_address(text):
	'''
	Test if a string is an valid email address
	'''
	regex = '^\w+([\.-]?\w+)*@\w+([\.-]?\w+)*(\.\w{2,3})+$'
	# and the string in search() method 
	return re.match(regex,text)

# Envelope class to keep in-memory details about a quilt patch
# This class is NOT modeled in the Database!
class QuiltPatch(object):
	def __init__(self, **kwargs):
		self.__dict__.update(kwargs)

	def __repr__(self):
		return self.__dict__

	def __str__(self):
		return str(self.__dict__)

def __review_debt(branch, debt):
	print("="*20, branch, "="*20)
	print("%-10s: %-5s (%-5s)" % ('domain', 'debt', 'derived'))
	print("-"*50)
	for d,plist in debt.items():
		print("%-10s: %-5d (%-5d)" % (d, sum(p.upstream_commit == '' for p in plist), sum(p.upstream_commit != '' for p in plist)))
#		for p in plist:
#			print("%-10s: %-40s %-40s %30s %s" % (d, p.commit_id, p.upstream_commit, p.upstream_author, p.subject))

def __sift_base(args):

	for kernel_name in args.kernel:
		kernel = Kernel.objects.get(name = kernel_name)
		base_version = KernelVersion(args.sift_base)
		csvfile = open("/tmp/bullpen_upstream_{}_{}.csv".format(args.sift_base,datetime.fromtimestamp(time.time()).strftime('%m%d')), "w")

		quilt_repo = KernelRepoSet.objects.get(kernel = kernel, repo__repotype__repotype = 'quilt').repo
		print(RGBColor(255,165,0, "\tProcessing Repo {}".format(quilt_repo.project)))
		__git = quilt_repo.initialize(scmdir=os.path.join(os.environ["WORKSPACE"],quilt_repo.project))	# returns a sh.Command() object
		ltsrepo = Repository.objects.get(repotype__repotype = 'lts')
		_ltsgit = ltsrepo.initialize(scmdir=os.path.join(os.environ["WORKSPACE"],ltsrepo.project))  # returns a sh.Command() object
		with pushd (quilt_repo.scmdir):
			# Use each tracking/staging branch configured for this kernel in each repo
			tbs = TrackerBranch.objects.filter(kernel = kernel, repo = quilt_repo)
			print(len(tbs))
			for tb in tbs:
				try:
					print(ANSIColor("bright_green", tb.branch))
					__git("checkout" , tb.branch)
				except Exception as e:
					print("Cannot checkout {}: {}".format(e.args[0].strip().split('\n')[-1]))
					continue
				debt = {}
				i = 0
				print('"PATCH FILE","ORIGINAL UPSTREAM TAG","MERGE LOCATION","UPSTREAM COMMIT","PAYLOAD HASH"', file = csvfile)
				with pushd("patches"):
					patchnames = open("series").read().splitlines()
					platform_tla = None
					extension = None
					for patchfile in patchnames:
						if not patchfile:
							print("BLANK LINE - continue")
							continue
						if patchfile[0] == '#':
							continue
						'''
						# FIXME - handle non-compliant lines in MLT
						m = re.search('^#[0-9a-f]+', patchfile)
						if m:
							continue
						# FIXME - handle non-compliant lines in MLT
						m = re.search('^ v5\..+', patchfile)
						if m:
							continue
						m = re.search('^#LABELS\s+P::\S+::(\S+)\s+D::(\S+)(?:\s+S::(\S+)){0,1}\s*.*$', patchfile)
						if m:
							platform_tla = m.group(1).upper()
							extension = m.group(2)
							continue		# Go to next line
						m = re.search('#LABELS\s+D::(\S+)', patchfile)
						if m:
							platform_tla = None
							extension = m.group(1)
							continue		# Go to next line
						# FIXME - handle non-compliant lines in Bullpen
						m = re.search('^#\s+(\w+)', patchfile)
						if m:
							extension = m.group(1)
							continue		# Go to next line
						basename, suffix = os.path.splitext(patchfile)
						# FIXME - handle non-compliant lines in MLT and Bullpen
						suffix = suffix[1:]
						# FIXME BEGIN
						if extension == 'thunderbolt':
							extension = 'tbt'
						if extension == 'usb-typec':
							extension = 'usbtypec'
						if extension == 'connectivity':
							extension = 'conn'
						if extension == 'turbostat':
							extension = 'trbostat'
						# FIXME END
						'''

#						if re.match('[0-9]{4}-revert-', patchfile.lower()):
#							continue
						with open(patchfile, errors="ignore") as infile:
							try:
								payload_hash = __git("patch-id", _in=infile, _tty_out=None).split()[0]
							except UnicodeDecodeError as e:
								print (ANSIColor("red", patchfile, ":", str(e)))
								continue
							except Exception as e:
								print (ANSIColor("red", str(e)))
								print (ANSIColor("red", patchfile, ":", "NO PATCH-ID found"))
								continue
						patches = StableKernel.objects.filter(payload_hash = payload_hash)
						if not patches:
							print(ANSIColor("yellow", patchfile, "NOT UPSTREAM"))
							continue
						highest_version = KernelVersion('v0.0')
						highest_patch = patches.first()
						for patch in patches:
							tag_version = KernelVersion(patch.tag)
							if tag_version > highest_version:
								highest_version = tag_version
								highest_patch = patch
						tag_version = highest_version
						patch = highest_patch
#					if tag_version <= base_version and ( tag_version.minor() == 0 or ( tag_version.major() == base_version.major() ) ) :
						with pushd(ltsrepo.scmdir):
							zzz = _ltsgit("rev-list", "--no-merges", "{}^..{}".format(patch.commit_id, args.sift_base), _tty_out=None).stdout.decode().strip().split('\n')
							if patch.commit_id in zzz:
								try:
									merge_location = _ltsgit("describe", "--contains", patch.commit_id, _tty_out=None).stdout.decode().strip()
								except Exception as e:
									merge_location = 'unavailable'
								print(ANSIColor("green", patchfile, "UPSTREAM AT", patch.tag, merge_location))
								print('"{}","{}","{}","{}","{}"'.format(patchfile, patch.tag, merge_location,patch.commit_id, payload_hash), file=csvfile)
							else:
								print(ANSIColor("red", patchfile, "UPSTREAM AT", patch.tag, "NOT REACHABLE FROM" , args.sift_base))
#					else:
#							print(patchfile, tag_version.major(), tag_version.minor(), base_version.major(), base_version.minor())
						i += 1
						if not (i % 100):
							print(i, "Patches Processed", file=sys.stderr)
					break


def __do_reports(args):

	for kernel_name in args.kernel:
		kernel = Kernel.objects.get(name = kernel_name)
		# FIXME monkey patch - don't call save() on this object!
		kernel.base_kernel = 'v'+kernel.base_kernel

		quilt_repo = KernelRepoSet.objects.get(kernel = kernel, repo__repotype__repotype = 'quilt').repo
		print(RGBColor(255,165,0, "\tProcessing Repo {}".format(quilt_repo.project)))
		__git = quilt_repo.initialize(scmdir=os.path.join(os.environ["WORKSPACE"],quilt_repo.project))	# returns a sh.Command() object
		with pushd (quilt_repo.scmdir):
			# Use each tracking/staging branch configured for this kernel in each repo
			tbs = TrackerBranch.objects.filter(kernel = kernel, repo = quilt_repo)
			print(len(tbs))
			for tb in tbs:
				try:
					print(ANSIColor("bright_green", tb.branch))
					__git("checkout" , tb.branch)
				except Exception as e:
					print("Cannot checkout {}: {}".format(e.args[0].strip().split('\n')[-1]))
					continue
				debt = {}
				i = 0
				with pushd("patches"):
					patchnames = open("series").read().splitlines()
					platform_tla = None
					extension = None
					for patchfile in patchnames:
						if not patchfile:
							print("BLANK LINE - continue")
							continue
						# FIXME - handle non-compliant lines in MLT
						m = re.search('^#[0-9a-f]+', patchfile)
						if m:
							continue
						# FIXME - handle non-compliant lines in MLT
						m = re.search('^ v5\..+', patchfile)
						if m:
							continue
						m = re.search('#LABELS\s+P::\w+::([,\w]+)\s+D::(\w+)\s*.*', patchfile)
						if m:
							platform_tla = m.group(1).upper()
							extension = m.group(2)
							continue		# Go to next line
						m = re.search('#LABELS\s+D::(\w+)', patchfile)
						if m:
							platform_tla = None
							extension = m.group(1)
							continue		# Go to next line
						# FIXME - handle non-compliant lines in Bullpen
						m = re.search('^#\s+(\w+)', patchfile)
						if m:
							extension = m.group(1)
							continue		# Go to next line
#						basename, extension = os.path.splitext(patchfile)
						basename, suffix = os.path.splitext(patchfile)
						# FIXME - handle non-compliant lines in MLT and Bullpen
						suffix = suffix[1:]
#						print(extension, suffix)
#						assert ( (suffix == extension) or ( suffix == 'patch' or suffix == 'EMBARGOED'))
						# FIXME BEGIN
						if extension == 'thunderbolt':
							extension = 'tbt'
						if extension == 'usb-typec':
							extension = 'usbtypec'
						if extension == 'connectivity':
							extension = 'conn'
						if extension == 'turbostat':
							extension = 'trbostat'
						# FIXME END

						if re.match('[0-9]{4}-revert-', patchfile.lower()):
							print (ANSIColor("cyan", patchfile, "is a REVERT - skip"))
							continue
						with open(patchfile, errors="ignore") as infile:
							try:
								payload_hash = __git("patch-id", _in=infile, _tty_out=None).split()[0]
							except UnicodeDecodeError as e:
								print (ANSIColor("red", patchfile, ":", str(e)))
								continue
							except Exception as e:
								print (ANSIColor("red", str(e)))
								print (ANSIColor("red", patchfile, ":", "NO PATCH-ID found"))
								continue
						patch = StableKernel.objects.filter(payload_hash = payload_hash).first()
						if not patch:
							patch_data = open(patchfile, "r").read()
							r = re.compile('^Subject:\s+(?:\[.+\])\s+(.+)$', flags=re.MULTILINE)
							subject = r.search(patch_data).group(1)
							r = re.compile('^From:\s+.*<(.+)>\s*$', flags=re.MULTILINE)
							try:
								author = r.search(patch_data).group(1)
							except:
								print(patch_data[:1000])
								sys.exit(1)
							r = re.compile('^From ([0-9a-f]+)\s+.+$', flags=re.MULTILINE)
							commit_id = r.search(patch_data).group(1)
							qp = QuiltPatch( commit_id = commit_id, subject = subject, author = author, upstream_commit = '', upstream_author = '', upstream_tag = '')
							if not platform_tla:
								platform_tla = ",".join(i[0].upper() for i in Platform.objects.filter(flags = 1).values_list('name'))
							qp.platform = platform_tla	# Monkey patch - DON'T do save() on this object
							patch = StableKernel.objects.filter(subject__startswith = subject ).first()
							if patch:
								print (ANSIColor("yellow", patchfile, "MAY BE DERIVED FROM", patch.subject, "commit", patch.commit_id , "tag", patch.tag))
								qp.upstream_commit = patch.commit_id
								qp.upstream_author = patch.author
								qp.subject = patch.subject	# in case of quilt truncation 
								qp.upstream_tag = patch.tag
							else:
								if not __intel_authored(author):
									print (ANSIColor("cyan", patchfile, extension, payload_hash, "NON-INTEL-AUTHORED", author))
									continue
								else:
									print (ANSIColor("magenta", patchfile, extension, payload_hash, qp.platform, "NOT UPSTREAM"))
							if not extension in debt:
								debt[extension] = []
							debt[extension].append(qp)
						else:
							if args.verbose:
								print(patchfile, "HASH" , payload_hash, "UPSTREAM at" , ANSIColor("yellow", patch.tag), "commit", patch.commit_id)
							pass
						i += 1
						if not (i % 100):
							print(i, "Patches Processed")
				__review_debt(tb.branch, debt)
				if args.report:
					__compute_domain_report_xlsx(kernel, debt, tb.branch)

def __non_upstream_len(l:list):
	count = 0
	for i in l:
		if not i.upstream_commit:
			count += 1
	return count

def __compute_domain_report_xlsx(kernel, debt, staging_branch):
	'''
	Produce a Microsoft Excel 201x workbook (.XLSX) file
	breaking down Technical Debt By Domain
	'''
	from openpyxl import Workbook
	from openpyxl.styles import Font
	from openpyxl.styles.fonts import DEFAULT_FONT

	def __size_columns(ws):
		for column_cells in ws.columns:
			cl = column_cells[0].column_letter
			for cell in column_cells:
				if cell.value is not None:
					current_width = ws.column_dimensions[cl].width
					new_width = max(current_width, len(str(cell.value)))
			new_width = (new_width + 2) * 1.2
			ws.column_dimensions[cl].width = new_width

	wb = Workbook()
	# NOTE: using a monospace font helps to properly auto-size the columns
	DEFAULT_FONT.name = 'Courier New'

	print("Calculating Debt By Domain ... ")

	for dd in debt.keys():
		print(dd)
		domain = Domain.objects.get(label_name = dd)
		ws = wb.create_sheet(re.sub('/','_',domain.name))
		row=1
		column=1
		ws.cell(row=row,column=column, value="PLATFORM")
		column += 1
		ws.cell(row=row,column=column, value="DOMAIN")
		column += 1
		ws.cell(row=row,column=column, value="COMMIT ID")
		column += 1
		ws.cell(row=row,column=column, value="SUBJECT")
		column += 1
		ws.cell(row=row,column=column, value="AUTHOR")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM COMMIT")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM AUTHOR")
		column += 1
		ws.cell(row=row,column=column, value="UPSTREAM LOCATION")
		column += 1
		ws.cell(row=row,column=column, value="WHITELISTED BY")
		column += 1
		ws.cell(row=row,column=column, value="WHITELIST REASON")
		column = 1
		r = ws.row_dimensions[row]
		r.font = Font(bold=True)
		row += 1
		for dp in debt[dd]:
			ws.cell(row=row,column=column, value=dp.platform)
			column += 1
			ws.cell(row=row,column=column, value=domain.name)
			column += 1
			ws.cell(row=row,column=column, value=dp.commit_id)
			column += 1
			ws.cell(row=row,column=column, value=dp.subject)
			column += 1
			try:
				ws.cell(row=row,column=column, value=dp.author)
			except:
				try:
					print(dp.author)
					ws.cell(row=row,column=column, value=dp.author)
				except:
					print(dp)
					ws.cell(row=row,column=column, value='')
			column += 1
			ws.cell(row=row,column=column, value=dp.upstream_commit)
			column += 1
			ws.cell(row=row,column=column, value=dp.upstream_author)
			column += 1
			ws.cell(row=row,column=column, value=dp.upstream_tag)
			column += 1
			ws.cell(row=row,column=column, value="")
			column += 1
			ws.cell(row=row,column=column, value="")
			row += 1
			column = 1
		__size_columns(ws)

	ws = wb.create_sheet('Summary', 0)
	row=2
	column=1
	ws.cell(row=row,column=column, value="Kernel Technical Debt Summary")
	row += 1
	ws.cell(row=row,column=column, value="Staging Branch: {}".format(staging_branch))
	row += 2
	column = 1
	ws.cell(row=row,column=column, value="DOMAIN")
	column += 1
	ws.cell(row=row,column=column, value="COUNT")
	column = 1
	r = ws.row_dimensions[row]
	r.font = Font(bold=True)
	row += 1
	sumstartrow = row
	for dd in debt.keys():
		domain = Domain.objects.get(label_name = dd)
#		if domain.rfu_0 & 1:
#			continue
		ws.cell(row=row,column=column, value=domain.name)
		column += 1
		ws.cell(row=row,column=column, value=__non_upstream_len(debt[dd]))
		row += 1
		column = 1
	__size_columns(ws)

	sumendrow = row-1
	row += 1
	ws.cell(row=row,column=column, value="TOTAL")
	column += 1
	ws.cell(row=row,column=column, value="=SUM(B"+str(sumstartrow)+":B"+str(sumendrow))
	# Delete Default sheet named 'Sheet'
	del wb['Sheet']
	# Sort Sheets alphabetically by domain
	wb._sheets.sort(key=lambda ws: ws.title)
	# Move Summary sheet to front of sheet list
	summary_sheet = wb._sheets.pop(wb._sheets.index(wb['Summary']))
	wb._sheets.insert(0, summary_sheet)
	# Save Workbook
	printable_branch_name = re.sub('[\.\-\/]', '_', staging_branch)
	output_file = "/tmp/{}_{}_techdebt_{}.xlsx".format(kernel.current_baseline, printable_branch_name, datetime.fromtimestamp(time.time()).strftime('%m%d'))
	wb.save(output_file)
	print("Results saved to {}".format(output_file))
	sys.exit(0)

if __name__ == '__main__':
	parser = argparse.ArgumentParser(prog=sys.argv[0])
	parser.add_argument('--kernel', '-k', type=Kernel.validate, help="Valid Values: "+Kernel.list()+" default: all kernels in DB")
	parser.add_argument('--report', '-R', action='store_true', default=False, help="generate Excel report(s)")
	parser.add_argument('--verbose', '-v', action='store_true', default=False, help="generate full console output, including upstream patches")
	parser.add_argument('--sift_base', '-s', type=str, metavar="vX.Y[.Z]", help="Generate list of patches upstreamed before base kernel version")
	args = parser.parse_args()

	assert(os.path.exists(os.environ["WORKSPACE"]))

	if not args.kernel:
		args.kernel = json.loads(re.sub('\'','"',Kernel.list()))		# produce TD report for all supported kernels
	else:
		args.kernel = [ args.kernel ]
	print(args.kernel)
	if args.sift_base:
		__sift_base(args)
	else:
		__do_reports(args)
	print('Done')
